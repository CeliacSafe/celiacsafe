#!/usr/bin/env python3
"""
CeliacSafe — Commercial-links audit in Excel-Workbooks einspielen.

Liest ``verified_direct_links`` aus der Audit-Datei und aktualisiert
``delivery_links``, ``reservation_links`` sowie optionale Inline-URL-Spalten
im Blatt ``restaurants`` (thefork_url, uber_eats_url, …).

Restaurant-IDs aus dem Audit werden per ``restaurant_scope`` (Name + Stadt)
auf die IDs in den Ziel-Workbooks gemappt.
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("Fehler: openpyxl nicht installiert.")
    print("Bitte: pip install -r scripts/requirements.txt")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_AUDIT = (
    ROOT / "data-source" / "CeliacSafe_CommercialLinks_Audit_ES_DE_2026-06-09.xlsx"
)
DEFAULT_TARGETS = [
    ROOT / "data-source" / "CeliacSafe_Datenbank_v4_Spain_129_geocoded.xlsx",
    ROOT / "data-source" / "CeliacSafe_Datenbank_v4_Germany_Delta_Consolidated_geocoded.xlsx",
]

sys.path.insert(0, str(Path(__file__).resolve().parent))
from platform_urls import is_usable_url, normalize_platform  # noqa: E402


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def normalize_key(name: Any, city: Any) -> tuple[str, str]:
    def norm(part: Any) -> str:
        if part is None:
            return ""
        text = unicodedata.normalize("NFKD", str(part))
        text = text.encode("ascii", "ignore").decode("ascii").lower()
        return re.sub(r"[^a-z0-9]+", " ", text).strip()

    return norm(name), norm(city)


def read_sheet_rows(workbook: Any, sheet_name: str) -> tuple[list[dict[str, Any]], list[str]]:
    if sheet_name not in workbook.sheetnames:
        return [], []
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [normalize_header(cell) for cell in next(rows)]
    except StopIteration:
        return [], []
    records: list[dict[str, Any]] = []
    for values in rows:
        if not values or all(value in (None, "") for value in values):
            continue
        records.append(
            {
                headers[col_index]: values[col_index]
                for col_index in range(len(headers))
                if headers[col_index]
            }
        )
    return records, headers


def resolve_platform_code(platform: str, role: str) -> tuple[str | None, str | None]:
    raw = (platform or "").strip().lower()
    role_norm = (role or "").strip().lower()

    if raw == "own website":
        if role_norm == "delivery":
            return "own_delivery", "delivery"
        if role_norm == "reservation":
            return "own_website", "reservation"
        return None, None

    if raw == "reservator":
        return "own_website", "reservation"

    code = normalize_platform(platform)
    if not code:
        return None, None
    if role_norm == "delivery":
        return code, "delivery"
    if role_norm == "reservation":
        return code, "reservation"
    return None, None


INLINE_FIELD_BY_PLATFORM: dict[tuple[str, str], str] = {
    ("uber_eats", "delivery"): "uber_eats_url",
    ("glovo", "delivery"): "glovo_url",
    ("just_eat", "delivery"): "just_eat_url",
    ("wolt", "delivery"): "wolt_url",
    ("thefork", "reservation"): "thefork_url",
    ("own_delivery", "delivery"): "direct_order_url",
    ("own_website", "reservation"): "direct_reservation_url",
}

PLATFORM_LABELS: dict[str, str] = {
    "uber_eats": "Uber Eats",
    "glovo": "Glovo",
    "just_eat": "Just Eat",
    "wolt": "Wolt",
    "lieferando": "Lieferando",
    "thefork": "TheFork",
    "own_delivery": "Own delivery",
    "own_website": "Own website",
}


def build_scope_maps(audit_path: Path) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(audit_path, read_only=True, data_only=True)
    scope_rows, _ = read_sheet_rows(workbook, "restaurant_scope")
    verified_rows, _ = read_sheet_rows(workbook, "verified_direct_links")
    workbook.close()
    scope_by_id = {
        str(row["restaurant_id"]).strip(): row
        for row in scope_rows
        if row.get("restaurant_id")
    }
    return scope_by_id, verified_rows


def build_workbook_id_lookup(workbook_path: Path) -> dict[tuple[str, str], str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    restaurant_rows, _ = read_sheet_rows(workbook, "restaurants")
    workbook.close()
    lookup: dict[tuple[str, str], str] = {}
    for row in restaurant_rows:
        restaurant_id = row.get("id")
        if not restaurant_id:
            continue
        lookup[normalize_key(row.get("name"), row.get("city"))] = str(restaurant_id).strip()
    return lookup


def map_audit_id_to_workbook_id(
    audit_id: str,
    scope_by_id: dict[str, dict[str, Any]],
    workbook_lookup: dict[tuple[str, str], str],
) -> str | None:
    if audit_id in workbook_lookup.values():
        return audit_id

    scope_row = scope_by_id.get(audit_id)
    if not scope_row:
        return None
    return workbook_lookup.get(normalize_key(scope_row.get("name"), scope_row.get("city")))


def collect_applicable_links(
    verified_rows: list[dict[str, Any]],
    scope_by_id: dict[str, dict[str, Any]],
    workbook_lookup: dict[tuple[str, str], str],
) -> list[dict[str, Any]]:
    applicable: list[dict[str, Any]] = []

    for row in verified_rows:
        if str(row.get("publish_cta", "")).strip().lower() != "yes":
            continue
        url = row.get("direct_url")
        if not is_usable_url(url):
            continue

        audit_id = str(row.get("restaurant_id", "")).strip()
        workbook_id = map_audit_id_to_workbook_id(audit_id, scope_by_id, workbook_lookup)
        if not workbook_id:
            continue

        platform_code, link_kind = resolve_platform_code(
            str(row.get("platform", "")),
            str(row.get("role", "")),
        )
        if not platform_code or not link_kind:
            continue

        scope_row = scope_by_id.get(audit_id, {})
        applicable.append(
            {
                "audit_id": audit_id,
                "workbook_id": workbook_id,
                "restaurant_name": row.get("restaurant_name")
                or scope_row.get("name")
                or workbook_id,
                "platform_code": platform_code,
                "link_kind": link_kind,
                "url": str(url).strip(),
                "link_status": row.get("link_status"),
                "source_url": row.get("source_url"),
                "note": row.get("note") or row.get("validation_note"),
                "last_checked": row.get("last_checked"),
            }
        )

    return applicable


def find_header_columns(headers: list[str]) -> dict[str, int]:
    return {header: index for index, header in enumerate(headers) if header}


def upsert_link_row(
    worksheet: Any,
    headers: list[str],
    *,
    link_id: str,
    restaurant_id: str,
    restaurant_name: str,
    platform_label: str,
    url: str,
    last_checked: Any,
    notes: str | None,
    row_index_by_key: dict[tuple[str, str], int],
) -> bool:
    header_cols = find_header_columns(headers)
    platform_col = header_cols.get("platform")
    url_col = header_cols.get("url")
    restaurant_id_col = header_cols.get("restaurant_id")
    if platform_col is None or url_col is None or restaurant_id_col is None:
        return False

    platform_code = normalize_platform(platform_label) or platform_label
    existing_row = row_index_by_key.get((restaurant_id, platform_code))

    if existing_row is None:
        existing_row = worksheet.max_row + 1
        row_index_by_key[(restaurant_id, platform_code)] = existing_row
        if "id" in header_cols:
            worksheet.cell(row=existing_row, column=header_cols["id"] + 1, value=link_id)
        if "restaurant_name" in header_cols:
            worksheet.cell(
                row=existing_row,
                column=header_cols["restaurant_name"] + 1,
                value=restaurant_name,
            )

    worksheet.cell(row=existing_row, column=restaurant_id_col + 1, value=restaurant_id)
    worksheet.cell(row=existing_row, column=platform_col + 1, value=platform_label)
    worksheet.cell(row=existing_row, column=url_col + 1, value=url)

    if "is_active" in header_cols:
        worksheet.cell(row=existing_row, column=header_cols["is_active"] + 1, value=True)
    if last_checked and "last_checked" in header_cols:
        worksheet.cell(row=existing_row, column=header_cols["last_checked"] + 1, value=last_checked)
    if notes and "notes" in header_cols:
        worksheet.cell(row=existing_row, column=header_cols["notes"] + 1, value=notes)
    if "verification_status" in header_cols:
        worksheet.cell(
            row=existing_row,
            column=header_cols["verification_status"] + 1,
            value="verified_from_commercial_links_audit",
        )
    if "verification_source_url" in header_cols and notes:
        worksheet.cell(
            row=existing_row,
            column=header_cols["verification_source_url"] + 1,
            value=notes,
        )

    return existing_row == worksheet.max_row


def apply_to_workbook(
    workbook_path: Path,
    applicable_links: list[dict[str, Any]],
) -> dict[str, int]:
    workbook = load_workbook(workbook_path)
    stats = {
        "matched": 0,
        "inline_updates": 0,
        "delivery_updates": 0,
        "reservation_updates": 0,
        "skipped": 0,
    }

    if "restaurants" not in workbook.sheetnames:
        workbook.close()
        return stats

    restaurant_sheet = workbook["restaurants"]
    header_row = next(restaurant_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    restaurant_headers = [normalize_header(cell) for cell in header_row]
    restaurant_header_cols = find_header_columns(restaurant_headers)

    if "id" not in restaurant_header_cols:
        workbook.close()
        raise ValueError(f"'id'-Spalte fehlt in {workbook_path.name}")

    id_col = restaurant_header_cols["id"] + 1
    row_by_id: dict[str, int] = {}
    for row_index, values in enumerate(
        restaurant_sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not values:
            continue
        restaurant_id = values[restaurant_header_cols["id"]]
        if restaurant_id:
            row_by_id[str(restaurant_id).strip()] = row_index

    workbook_ids = set(row_by_id)
    links_for_workbook = [link for link in applicable_links if link["workbook_id"] in workbook_ids]
    stats["matched"] = len(links_for_workbook)

    for link in links_for_workbook:
        row_index = row_by_id.get(link["workbook_id"])
        if row_index is None:
            stats["skipped"] += 1
            continue

        inline_field = INLINE_FIELD_BY_PLATFORM.get(
            (link["platform_code"], link["link_kind"])
        )
        if inline_field and inline_field in restaurant_header_cols:
            restaurant_sheet.cell(
                row=row_index,
                column=restaurant_header_cols[inline_field] + 1,
                value=link["url"],
            )
            stats["inline_updates"] += 1

        if "commercial_link_last_checked" in restaurant_header_cols and link.get("last_checked"):
            restaurant_sheet.cell(
                row=row_index,
                column=restaurant_header_cols["commercial_link_last_checked"] + 1,
                value=link["last_checked"],
            )

    for sheet_name, stat_key, prefix in (
        ("delivery_links", "delivery_updates", "DL"),
        ("reservation_links", "reservation_updates", "RS"),
    ):
        if sheet_name not in workbook.sheetnames:
            continue

        worksheet = workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [normalize_header(cell) for cell in header_row]
        header_cols = find_header_columns(headers)
        if "restaurant_id" not in header_cols or "platform" not in header_cols:
            continue

        row_index_by_key: dict[tuple[str, str], int] = {}
        for row_index in range(2, worksheet.max_row + 1):
            restaurant_id = worksheet.cell(
                row=row_index, column=header_cols["restaurant_id"] + 1
            ).value
            platform_raw = worksheet.cell(row=row_index, column=header_cols["platform"] + 1).value
            platform_code = normalize_platform(platform_raw)
            if restaurant_id and platform_code:
                row_index_by_key[(str(restaurant_id).strip(), platform_code)] = row_index

        for link in links_for_workbook:
            if link["link_kind"] != sheet_name.replace("_links", ""):
                continue

            platform_label = PLATFORM_LABELS.get(link["platform_code"], link["platform_code"])
            link_id = f"{prefix}-{link['workbook_id']}-{link['platform_code'].upper()}"
            created = upsert_link_row(
                worksheet,
                headers,
                link_id=link_id,
                restaurant_id=link["workbook_id"],
                restaurant_name=str(link["restaurant_name"]),
                platform_label=platform_label,
                url=link["url"],
                last_checked=link.get("last_checked"),
                notes=link.get("note"),
                row_index_by_key=row_index_by_key,
            )
            if created:
                stats[stat_key] += 1
            else:
                stats[stat_key] += 1

    workbook.save(workbook_path)
    workbook.close()
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description="Spielt Commercial-Links-Audit in Excel ein.")
    parser.add_argument("-i", "--input", type=Path, default=DEFAULT_AUDIT)
    parser.add_argument(
        "-t",
        "--target",
        type=Path,
        action="append",
        dest="targets",
        help="Ziel-Workbook (mehrfach moeglich)",
    )
    args = parser.parse_args()

    targets = args.targets or DEFAULT_TARGETS
    if not args.input.exists():
        print(f"Fehler: Audit-Datei nicht gefunden: {args.input}")
        sys.exit(1)

    scope_by_id, verified_rows = build_scope_maps(args.input)
    publish_rows = [
        row
        for row in verified_rows
        if str(row.get("publish_cta", "")).strip().lower() == "yes"
        and is_usable_url(row.get("direct_url"))
    ]
    print(f"Audit: {len(publish_rows)} verifizierte Links mit publish_cta=yes")

    for target in targets:
        if not target.exists():
            print(f"Warnung: Ziel nicht gefunden, uebersprungen: {target}")
            continue

        workbook_lookup = build_workbook_id_lookup(target)
        applicable = collect_applicable_links(verified_rows, scope_by_id, workbook_lookup)
        stats = apply_to_workbook(target, applicable)
        print(f"\n{target.name}")
        print(f"  Gemappte Links: {stats['matched']}")
        print(f"  Inline-URL-Updates: {stats['inline_updates']}")
        print(f"  Delivery-Links: {stats['delivery_updates']}")
        print(f"  Reservierungs-Links: {stats['reservation_updates']}")


if __name__ == "__main__":
    main()
