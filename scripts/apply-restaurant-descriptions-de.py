#!/usr/bin/env python3
"""
CeliacSafe — Deutsche Restaurantbeschreibungen aus Excel einspielen.

Liest ``restaurant_descriptions`` (Spalten: Restaurantname, Beschreibung)
und schreibt ``description_de`` ins Blatt ``restaurants``.

Zuordnung per normalisiertem Restaurantnamen (eindeutig über ES+DE-Scope).
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("Fehler: openpyxl nicht installiert.")
    print("Bitte: pip install -r scripts/requirements.txt")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE = (
    ROOT / "data-source" / "CeliacSafe_Restaurantbeschreibungen_ES_DE_181.xlsx"
)
DEFAULT_TARGETS = [
    ROOT / "data-source" / "CeliacSafe_Datenbank_v4_Spain_129_geocoded.xlsx",
    ROOT / "data-source" / "CeliacSafe_Datenbank_v4_Germany_Delta_Consolidated_geocoded.xlsx",
]

NAME_ALIASES: dict[str, str] = {
    "restaurantname": "name",
    "beschreibung": "description_de",
}


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    raw = str(value).strip().lower().replace(" ", "_")
    return NAME_ALIASES.get(raw, raw)


def normalize_name(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = text.encode("ascii", "ignore").decode("ascii").lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def parse_string(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def read_sheet_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [normalize_header(cell) for cell in next(rows)]
    except StopIteration:
        return []
    records: list[dict[str, Any]] = []
    for values in rows:
        if not values or all(value in (None, "") for value in values):
            continue
        records.append(
            {
                headers[col_index]: values[col_index]
                for col_index in range(len(headers))
                if headers[col_index]
            }
        )
    return records


def load_descriptions_by_name(source_path: Path) -> dict[str, str]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    rows = read_sheet_rows(workbook, "restaurant_descriptions")
    workbook.close()

    descriptions: dict[str, str] = {}
    for row in rows:
        name = parse_string(row.get("name"))
        description = parse_string(row.get("description_de"))
        if not name or not description:
            continue
        key = normalize_name(name)
        descriptions[key] = description
    return descriptions


def build_workbook_name_index(workbook_path: Path) -> dict[str, str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    restaurant_rows = read_sheet_rows(workbook, "restaurants")
    workbook.close()

    index: dict[str, str] = {}
    for row in restaurant_rows:
        restaurant_id = row.get("id")
        name = row.get("name")
        if not restaurant_id or not name:
            continue
        index[normalize_name(name)] = str(restaurant_id).strip()
    return index


def collect_updates(
    descriptions_by_name: dict[str, str],
    name_index: dict[str, str],
) -> dict[str, str]:
    updates: dict[str, str] = {}
    for name_key, description in descriptions_by_name.items():
        restaurant_id = name_index.get(name_key)
        if restaurant_id:
            updates[restaurant_id] = description
    return updates


def apply_to_workbook(
    workbook_path: Path,
    updates_by_id: dict[str, str],
) -> dict[str, int]:
    workbook = load_workbook(workbook_path)
    stats = {"matched": 0, "description_de": 0}

    if "restaurants" not in workbook.sheetnames:
        workbook.close()
        return stats

    sheet = workbook["restaurants"]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(cell) for cell in header_row]
    header_cols = {header: index for index, header in enumerate(headers) if header}

    if "id" not in header_cols or "description_de" not in header_cols:
        workbook.close()
        raise ValueError(f"'id' oder 'description_de' fehlt in {workbook_path.name}")

    for row_index, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not values:
            continue
        restaurant_id = values[header_cols["id"]]
        if not restaurant_id:
            continue
        restaurant_id = str(restaurant_id).strip()
        description = updates_by_id.get(restaurant_id)
        if not description:
            continue

        stats["matched"] += 1
        sheet.cell(row=row_index, column=header_cols["description_de"] + 1, value=description)
        stats["description_de"] += 1

    workbook.save(workbook_path)
    workbook.close()
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Spielt deutsche Restaurantbeschreibungen in Excel ein."
    )
    parser.add_argument("-i", "--input", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument(
        "-t",
        "--target",
        type=Path,
        action="append",
        dest="targets",
        help="Ziel-Workbook (mehrfach moeglich)",
    )
    args = parser.parse_args()

    targets = args.targets or DEFAULT_TARGETS
    if not args.input.exists():
        print(f"Fehler: Quelldatei nicht gefunden: {args.input}")
        sys.exit(1)

    descriptions_by_name = load_descriptions_by_name(args.input)
    print(f"Quelle: {len(descriptions_by_name)} Beschreibungen")

    for target in targets:
        if not target.exists():
            print(f"Warnung: Ziel nicht gefunden, uebersprungen: {target}")
            continue

        name_index = build_workbook_name_index(target)
        updates = collect_updates(descriptions_by_name, name_index)
        stats = apply_to_workbook(target, updates)
        print(f"\n{target.name}")
        print(f"  Gemappte Restaurants: {stats['matched']}")
        print(f"  description_de: {stats['description_de']}")


if __name__ == "__main__":
    main()
