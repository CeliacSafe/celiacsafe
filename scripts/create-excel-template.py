#!/usr/bin/env python3
"""
Erzeugt eine leere CeliacSafe-Excel-Vorlage mit allen Blättern und Kopfzeilen.

Verwendung:
    python scripts/create-excel-template.py
    python scripts/create-excel-template.py -o data-source/Meine_Datei.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT = ROOT / "data-source" / "CeliacSafe_Datenbank_v4.xlsx"

SHEETS: dict[str, list[str]] = {
    "README": [],  # filled as prose, not column headers
    "lookups": [
        "country_code",
        "country_name",
        "region_code_es",
        "region_name_es",
        "venue_type",
        "cuisine_type",
        "price_range",
        "meal_type",
        "verification_status",
        "verification_method",
        "platform_delivery",
        "platform_reservation",
        "data_source_type",
        "national_authority",
    ],
    "restaurants": [
        "id",
        "name",
        "slug",
        "country_code",
        "region_code",
        "region_name",
        "province",
        "city",
        "district",
        "postal_code",
        "address_street",
        "latitude",
        "longitude",
        "google_maps_url",
        "apple_maps_url",
        "venue_type",
        "cuisine_types",
        "price_range",
        "meal_types",
        "verification_status",
        "verification_methods",
        "last_verified_at",
        "face_program",
        "aoecs_certified",
        "national_authority",
        "phone",
        "whatsapp",
        "email",
        "website",
        "menu_url",
        "instagram",
        "facebook",
        "opening_hours",
        "seasonal_closure",
        "description_de",
        "description_en",
        "description_es",
        "featured_image_url",
        "is_published",
        "is_hidden",
        "thefork_url",
        "glovo_url",
        "uber_eats_url",
        "just_eat_url",
    ],
    "delivery_links": [
        "id",
        "restaurant_id",
        "restaurant_name",
        "platform",
        "url",
        "is_active",
        "last_checked",
        "notes",
    ],
    "reservation_links": [
        "id",
        "restaurant_id",
        "restaurant_name",
        "platform",
        "url",
        "is_active",
        "last_checked",
        "notes",
    ],
    "submissions": [
        "id",
        "restaurant_name",
        "city",
        "country_code",
        "address",
        "website",
        "phone",
        "submission_notes",
        "submitted_by_email",
        "submitted_by_name",
        "submission_status",
        "source",
        "submitted_at",
    ],
    "statistics": [
        "metric",
        "value",
        "notes",
    ],
}

README_LINES = [
    "CeliacSafe — Master-Datenbank (leere Vorlage)",
    "",
    "Spaltenköpfe: Zeile 1, snake_case. Daten ab Zeile 2.",
    "Doku: docs/EXCEL-SPALTENLISTE.md · docs/EXCEL-AUFBAU-CLAUDE.md",
    "",
    "Blätter:",
    "  lookups            — erlaubte Werte (Referenz)",
    "  restaurants        — ein Lokal pro Zeile (Pflicht für Export)",
    "  delivery_links     — Lieferplattformen (0–n pro Restaurant)",
    "  reservation_links  — Reservierung (0–n pro Restaurant)",
    "  submissions        — Community-Vorschläge (optional)",
    "  statistics         — Zähler / Notizen (optional)",
    "",
    "Export: npm run data:build",
    "Web-Admin: admin-web (Supabase)",
]

LOOKUP_SEEDS: dict[str, list[str]] = {
    "country_code": ["ES"],
    "country_name": ["España"],
    "region_code_es": [
        "ES-AN",
        "ES-AR",
        "ES-AS",
        "ES-CN",
        "ES-CB",
        "ES-CL",
        "ES-CM",
        "ES-CT",
        "ES-EX",
        "ES-GA",
        "ES-IB",
        "ES-RI",
        "ES-MD",
        "ES-MC",
        "ES-NC",
        "ES-PV",
        "ES-VC",
        "ES-CE",
        "ES-ML",
    ],
    "venue_type": [
        "restaurant",
        "cafe",
        "bakery",
        "pastry_shop",
        "ice_cream",
        "pizzeria",
        "bar_tapas",
        "fast_food",
        "hotel_restaurant",
        "food_truck",
        "catering",
        "brunch_place",
        "burger_joint",
        "asian_restaurant",
    ],
    "price_range": ["€", "€€", "€€€", "€€€€"],
    "meal_type": ["breakfast", "brunch", "lunch", "dinner", "snacks", "drinks"],
    "verification_status": [
        "to_be_verified",
        "pending_verification",
        "in_verification",
        "verified",
        "rejected",
    ],
    "verification_method": [
        "own_visit",
        "phone_confirmed",
        "email_confirmed",
        "face_certified",
        "regional_assoc_certified",
        "operator_declaration",
        "multiple_sources",
    ],
    "platform_delivery": [
        "glovo",
        "uber_eats",
        "just_eat",
        "wolt",
        "lieferando",
        "foodora",
        "takeaway",
        "bolt_food",
        "own_delivery",
        "no_delivery",
    ],
    "platform_reservation": [
        "thefork",
        "opentable",
        "quandoo",
        "booking_restaurants",
        "own_website",
        "phone_only",
        "walk_in_only",
        "instagram_dm",
    ],
}

HEADER_FONT = Font(bold=True)


def write_header_row(ws: Worksheet, headers: list[str]) -> None:
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(28, len(name) + 2))


def fill_lookups(ws: Worksheet, headers: list[str]) -> None:
    write_header_row(ws, headers)
    col_by_name = {name: idx + 1 for idx, name in enumerate(headers)}
    max_rows = max(len(v) for v in LOOKUP_SEEDS.values()) if LOOKUP_SEEDS else 0
    for row_offset in range(max_rows):
        row = row_offset + 2
        for key, values in LOOKUP_SEEDS.items():
            col = col_by_name.get(key)
            if col is None or row_offset >= len(values):
                continue
            ws.cell(row=row, column=col, value=values[row_offset])


def fill_readme(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 72
    for row, line in enumerate(README_LINES, start=1):
        cell = ws.cell(row=row, column=1, value=line)
        if row == 1:
            cell.font = Font(bold=True, size=12)


def build_workbook() -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for name in [
        "README",
        "lookups",
        "restaurants",
        "delivery_links",
        "reservation_links",
        "submissions",
        "statistics",
    ]:
        ws = wb.create_sheet(title=name)
        headers = SHEETS[name]
        if name == "README":
            fill_readme(ws)
        elif name == "lookups":
            fill_lookups(ws, headers)
        else:
            write_header_row(ws, headers)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="Leere CeliacSafe-Excel-Vorlage erzeugen")
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(args.output)
    print(f"Erstellt: {args.output}")
    print(f"Blätter: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
