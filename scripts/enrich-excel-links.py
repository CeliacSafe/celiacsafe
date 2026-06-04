#!/usr/bin/env python3
"""
Schreibt fehlende Plattform-URLs in die Excel-Datei zurueck.

Aktualisiert die Spalte ``url`` in den Blaettern ``delivery_links`` und
``reservation_links``, wenn die Zelle leer ist und die Plattform unterstuetzt wird.

Verwendung::

    python scripts/enrich-excel-links.py
    python scripts/enrich-excel-links.py --output data-source/CeliacSafe_Datenbank_v4_enriched.xlsx
    python scripts/enrich-excel-links.py --in-place
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("Fehler: openpyxl nicht installiert.")
    print("Bitte: pip install -r scripts/requirements.txt")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = ROOT / "data-source" / "CeliacSafe_Datenbank_v4.xlsx"

# Import aus gleichem Ordner
sys.path.insert(0, str(Path(__file__).resolve().parent))
from platform_urls import (  # noqa: E402
    build_url_for_platform,
    is_usable_url,
    normalize_platform,
)


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def read_restaurant_lookup(path: Path) -> dict[str, dict[str, Any]]:
    """Minimaler Restaurant-Lookup aus dem Blatt restaurants."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows, _ = _read_sheet_rows(workbook, "restaurants")
    workbook.close()

    lookup: dict[str, dict[str, Any]] = {}
    for row in rows:
        rid = row.get("id")
        if rid is None:
            continue
        lookup[str(rid).strip()] = {
            "id": str(rid).strip(),
            "name": row.get("name"),
            "city": row.get("city"),
            "slug": row.get("slug"),
            "website": row.get("website"),
        }
    return lookup


def _read_sheet_rows(workbook: Any, sheet_name: str) -> tuple[list[dict[str, Any]], int]:
    if sheet_name not in workbook.sheetnames:
        return [], 0
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [normalize_header(cell) for cell in next(rows)]
    except StopIteration:
        return [], 0
    records: list[dict[str, Any]] = []
    for values in rows:
        if not values or all(value in (None, "") for value in values):
            continue
        records.append(
            {
                headers[col_index]: values[col_index]
                for col_index in range(len(headers))
                if headers[col_index]
            }
        )
    return records, len(headers)


def enrich_link_sheet(
    workbook: Any,
    sheet_name: str,
    restaurant_lookup: dict[str, dict[str, Any]],
) -> int:
    if sheet_name not in workbook.sheetnames:
        return 0

    worksheet = workbook[sheet_name]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(cell) for cell in header_row]

    try:
        platform_col = headers.index("platform") + 1
        url_col = headers.index("url") + 1
        restaurant_id_col = headers.index("restaurant_id") + 1
    except ValueError:
        print(f"Warnung: Spalten in '{sheet_name}' unvollstaendig — uebersprungen.")
        return 0

    updated = 0
    for row_idx in range(2, worksheet.max_row + 1):
        platform_raw = worksheet.cell(row=row_idx, column=platform_col).value
        platform = normalize_platform(platform_raw)
        if not platform:
            continue

        url_cell = worksheet.cell(row=row_idx, column=url_col)
        if is_usable_url(url_cell.value):
            continue

        restaurant_id = worksheet.cell(row=row_idx, column=restaurant_id_col).value
        if restaurant_id is None:
            continue

        restaurant = restaurant_lookup.get(str(restaurant_id).strip())
        if not restaurant:
            continue

        built = build_url_for_platform(platform, restaurant, None)
        if not built:
            continue

        url_cell.value = built
        updated += 1

    return updated


def main() -> None:
    parser = argparse.ArgumentParser(description="Ergaenzt leere Plattform-URLs in der Excel-Datei.")
    parser.add_argument(
        "--input",
        "-i",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Eingabe-xlsx (Standard: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=None,
        help="Ausgabe-xlsx (Standard: Eingabe mit Suffix _enriched)",
    )
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="Ueberschreibt die Eingabedatei",
    )
    args = parser.parse_args()

    input_path = args.input
    if not input_path.exists():
        print(f"Fehler: Datei nicht gefunden: {input_path}")
        sys.exit(1)

    if args.in_place:
        output_path = input_path
    else:
        output_path = args.output or input_path.with_stem(input_path.stem + "_enriched")

    print(f"Lese: {input_path}")
    restaurant_lookup = read_restaurant_lookup(input_path)

    workbook = load_workbook(input_path)
    delivery_updated = enrich_link_sheet(workbook, "delivery_links", restaurant_lookup)
    reservation_updated = enrich_link_sheet(workbook, "reservation_links", restaurant_lookup)

    workbook.save(output_path)
    workbook.close()

    print(f"Gespeichert: {output_path}")
    print(f"Aktualisierte Zeilen: {delivery_updated} Lieferung, {reservation_updated} Reservierung")


if __name__ == "__main__":
    main()
