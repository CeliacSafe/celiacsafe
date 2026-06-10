#!/usr/bin/env python3
"""
CeliacSafe — Germany delta Excel → v4 restaurants workbook.

Reads ``delta_publish_all`` from a consolidated Germany delta file and writes
a standard ``restaurants`` sheet compatible with ``excel-to-json.py`` / geocode.

Optionally merges seed-only rows from the previous Germany seed file.
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from copy import copy
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("Fehler: openpyxl nicht installiert.")
    print("Bitte ausfuehren: pip install -r scripts/requirements.txt")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_DELTA = (
    ROOT / "data-source" / "CeliacSafe_Delta_DE_Complete_Consolidated.xlsx"
)
DEFAULT_SEED = (
    ROOT / "data-source" / "CeliacSafe_Datenbank_v4_Germany_100GF_Geocoded_Seed.xlsx"
)
DEFAULT_OUTPUT = (
    ROOT / "data-source" / "CeliacSafe_Datenbank_v4_Germany_Delta_Consolidated.xlsx"
)

PUBLISH_ACTIONS = {
    "publish",
    "publish_after_direct_confirm_or_soft_publish",
}

DE_STATE_TO_REGION: dict[str, tuple[str, str, str]] = {
    "Hessen": ("DE-HE", "Hessen", "Hessen"),
    "Berlin": ("DE-BE", "Berlin", "Berlin"),
    "NRW": ("DE-NW", "North Rhine-Westphalia", "Nordrhein-Westfalen"),
    "Nordrhein-Westfalen": ("DE-NW", "North Rhine-Westphalia", "Nordrhein-Westfalen"),
    "Baden-Württemberg": ("DE-BW", "Baden-Württemberg", "Baden-Württemberg"),
    "Bayern": ("DE-BY", "Bavaria", "Bayern"),
    "Hamburg": ("DE-HH", "Hamburg", "Hamburg"),
    "Schleswig-Holstein": ("DE-SH", "Schleswig-Holstein", "Schleswig-Holstein"),
    "Thüringen": ("DE-TH", "Thuringia", "Thüringen"),
    "Rheinland-Pfalz": ("DE-RP", "Rhineland-Palatinate", "Rheinland-Pfalz"),
    "Niedersachsen": ("DE-NI", "Lower Saxony", "Niedersachsen"),
    "Sachsen": ("DE-SN", "Saxony", "Sachsen"),
    "Sachsen-Anhalt": ("DE-ST", "Saxony-Anhalt", "Sachsen-Anhalt"),
    "Brandenburg": ("DE-BB", "Brandenburg", "Brandenburg"),
    "Bremen": ("DE-HB", "Bremen", "Bremen"),
    "Saarland": ("DE-SL", "Saarland", "Saarland"),
    "Mecklenburg-Vorpommern": (
        "DE-MV",
        "Mecklenburg-Western Pomerania",
        "Mecklenburg-Vorpommern",
    ),
}


def parse_string(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = re.sub(r"[^\w\s-]", "", ascii_text.lower())
    slug = re.sub(r"[-\s]+", "-", ascii_text).strip("-")
    return slug[:80] or "restaurant"


def normalize_key(name: str | None, city: str | None) -> tuple[str, str]:
    def norm(part: str | None) -> str:
        if not part:
            return ""
        text = unicodedata.normalize("NFKD", part)
        text = text.encode("ascii", "ignore").decode("ascii").lower()
        return re.sub(r"[^a-z0-9]+", " ", text).strip()

    return norm(name), norm(city)


def resolve_region(federal_state: str | None) -> tuple[str, str, str]:
    if not federal_state:
        return "DE", "Germany", "Deutschland"
    key = federal_state.strip()
    if key in DE_STATE_TO_REGION:
        return DE_STATE_TO_REGION[key]
    return "DE", key, key


def category_to_cuisine(category: str | None) -> str | None:
    if not category:
        return None
    parts = re.split(r"[/,]", category)
    cleaned = [part.strip() for part in parts if part.strip()]
    return ", ".join(cleaned) if cleaned else None


def delta_row_to_restaurant(row: dict[str, Any]) -> dict[str, Any]:
    region_code, region_name, province = resolve_region(parse_string(row.get("federal_state")))
    name = parse_string(row.get("name")) or ""
    city = parse_string(row.get("city")) or ""
    restaurant_id = parse_string(row.get("delta_row_id")) or parse_string(row.get("restaurant_id"))
    if not restaurant_id:
        raise ValueError(f"Restaurant ohne ID: {name}")

    record: dict[str, Any] = {
        "id": restaurant_id,
        "name": name,
        "slug": slugify(f"{name}-{city}"),
        "country_code": "DE",
        "region_code": region_code,
        "region_name": region_name,
        "province": province,
        "city": city,
        "district": parse_string(row.get("area")),
        "postal_code": parse_string(row.get("postal_code")),
        "address_street": parse_string(row.get("street_address")),
        "google_maps_deeplink": parse_string(row.get("google_maps_deeplink")),
        "apple_maps_deeplink": parse_string(row.get("apple_maps_deeplink")),
        "venue_type": parse_string(row.get("category")),
        "cuisine_types": category_to_cuisine(parse_string(row.get("category"))),
        "verification_status": parse_string(row.get("verification_status"))
        or "verified_100_percent_gluten_free",
        "verification_methods": "multiple_sources",
        "last_verified_at": parse_string(row.get("last_checked")),
        "description_de": None,
        "description_en": None,
        "description_es": None,
        "is_published": True,
        "is_hidden": False,
        "verification_source": parse_string(row.get("verification_source")),
        "verification_note": parse_string(row.get("verification_note")),
        "verification_source_url": parse_string(row.get("source_urls")),
        "geo_latitude": row.get("geo_latitude"),
        "geo_longitude": row.get("geo_longitude"),
        "geocode_quality": parse_string(row.get("geocode_quality")),
        "geocode_status": parse_string(row.get("geocode_status")),
        "geocoding_query": ", ".join(
            part
            for part in [
                parse_string(row.get("street_address")),
                parse_string(row.get("postal_code")),
                parse_string(row.get("city")),
                "Germany",
            ]
            if part
        ),
    }

    delivery_url = parse_string(row.get("delivery_url"))
    reservation_url = parse_string(row.get("reservation_url"))
    if delivery_url:
        record["_delivery_url"] = delivery_url
    if reservation_url:
        record["_reservation_url"] = reservation_url

    return record


def read_delta_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "delta_publish_all" not in workbook.sheetnames:
        workbook.close()
        raise ValueError("Blatt 'delta_publish_all' nicht gefunden.")

    sheet = workbook["delta_publish_all"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    records: list[dict[str, Any]] = []

    for values in rows:
        if not values or not any(values):
            continue
        row = {
            headers[index]: values[index]
            for index in range(len(headers))
            if headers[index]
        }
        action = parse_string(row.get("import_action"))
        if action not in PUBLISH_ACTIONS:
            continue
        records.append(row)

    workbook.close()
    return records


def read_seed_restaurants(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "restaurants" not in workbook.sheetnames:
        workbook.close()
        return []

    sheet = workbook["restaurants"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    records: list[dict[str, Any]] = []

    for values in rows:
        if not values or not values[0]:
            continue
        record = {
            headers[index]: values[index]
            for index in range(len(headers))
            if headers[index]
        }
        records.append(record)

    workbook.close()
    return records


def merge_seed_only(
    delta_records: list[dict[str, Any]],
    seed_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    delta_keys = {
        normalize_key(parse_string(record.get("name")), parse_string(record.get("city")))
        for record in delta_records
    }
    merged = list(delta_records)
    seen_ids = {record["id"] for record in delta_records}

    for seed in seed_records:
        seed_id = parse_string(seed.get("id"))
        seed_key = normalize_key(parse_string(seed.get("name")), parse_string(seed.get("city")))
        if seed_key in delta_keys:
            continue
        if seed_id and seed_id in seen_ids:
            continue
        merged.append(seed)
        if seed_id:
            seen_ids.add(seed_id)

    return merged


def write_workbook(
    output_path: Path,
    restaurants: list[dict[str, Any]],
    template_headers: list[str],
) -> None:
    workbook = Workbook()
    restaurant_sheet = workbook.active
    restaurant_sheet.title = "restaurants"
    restaurant_sheet.append(template_headers)

    delivery_sheet = workbook.create_sheet("delivery_links")
    delivery_sheet.append(
        [
            "id",
            "restaurant_id",
            "restaurant_name",
            "platform",
            "url",
            "is_active",
            "last_checked",
            "notes",
        ]
    )

    reservation_sheet = workbook.create_sheet("reservation_links")
    reservation_sheet.append(
        [
            "id",
            "restaurant_id",
            "restaurant_name",
            "platform",
            "url",
            "is_active",
            "last_checked",
            "notes",
        ]
    )

    delivery_counter = 1
    reservation_counter = 1

    for restaurant in restaurants:
        row_values = [restaurant.get(header) for header in template_headers]
        restaurant_sheet.append(row_values)

        restaurant_id = parse_string(restaurant.get("id")) or ""
        restaurant_name = parse_string(restaurant.get("name")) or ""

        delivery_url = restaurant.pop("_delivery_url", None) if "_delivery_url" in restaurant else None
        reservation_url = (
            restaurant.pop("_reservation_url", None) if "_reservation_url" in restaurant else None
        )

        if delivery_url:
            delivery_sheet.append(
                [
                    f"DL-{delivery_counter:04d}",
                    restaurant_id,
                    restaurant_name,
                    "website",
                    delivery_url,
                    True,
                    restaurant.get("last_verified_at"),
                    "Imported from delta",
                ]
            )
            delivery_counter += 1

        if reservation_url:
            reservation_sheet.append(
                [
                    f"RS-{reservation_counter:04d}",
                    restaurant_id,
                    restaurant_name,
                    "website",
                    reservation_url,
                    True,
                    restaurant.get("last_verified_at"),
                    "Imported from delta",
                ]
            )
            reservation_counter += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def load_template_headers(seed_path: Path) -> list[str]:
    workbook = load_workbook(seed_path, read_only=True, data_only=True)
    headers = [
        str(cell).strip()
        for cell in next(workbook["restaurants"].iter_rows(min_row=1, max_row=1, values_only=True))
        if cell is not None and str(cell).strip()
    ]
    workbook.close()
    return headers


def main() -> None:
    parser = argparse.ArgumentParser(description="Konvertiert Germany-Delta-Excel nach v4-Format.")
    parser.add_argument("-i", "--input", type=Path, default=DEFAULT_DELTA)
    parser.add_argument(
        "--seed",
        type=Path,
        default=DEFAULT_SEED,
        help="Vorherige Germany-Seed-Datei fuer Merge seed-only Zeilen",
    )
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--no-seed-merge",
        action="store_true",
        help="Seed-only Zeilen nicht aus der alten Seed-Datei uebernehmen",
    )
    args = parser.parse_args()

    if not args.input.exists():
        print(f"Fehler: Delta-Datei nicht gefunden: {args.input}")
        sys.exit(1)

    delta_rows = read_delta_rows(args.input)
    delta_records = [delta_row_to_restaurant(row) for row in delta_rows]

    if args.no_seed_merge:
        final_records = delta_records
        merged_seed_count = 0
    else:
        seed_records = read_seed_restaurants(args.seed)
        final_records = merge_seed_only(delta_records, seed_records)
        merged_seed_count = len(final_records) - len(delta_records)

    template_headers = load_template_headers(args.seed)
    write_workbook(args.output, final_records, template_headers)

    with_coords = sum(
        1
        for record in final_records
        if record.get("geo_latitude") not in (None, "") and record.get("geo_longitude") not in (None, "")
    )

    print(f"Delta publish rows: {len(delta_records)}")
    print(f"Seed-only merged: {merged_seed_count}")
    print(f"Total restaurants: {len(final_records)}")
    print(f"With coordinates: {with_coords}")
    print(f"Output: {args.output}")


if __name__ == "__main__":
    main()
