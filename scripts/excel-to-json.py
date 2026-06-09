#!/usr/bin/env python3
"""
CeliacSafe — Excel-zu-JSON-Konverter.

Liest ``data-source/CeliacSafe_Datenbank_v3.xlsx`` und exportiert die Daten
nach ``src/data/restaurants.json`` im Format des TypeScript-Interfaces ``Restaurant``.

Blätter:
    - restaurants        → Hauptdaten (Zeile 1 = Spaltenköpfe)
    - delivery_links     → Liefer-Links, verknüpft über ``restaurant_id``
    - reservation_links  → Reservierungs-Links, verknüpft über ``restaurant_id``

Leere ``url``-Zellen werden beim Export automatisch ergänzt (TheFork, Glovo,
Uber Eats, …) — siehe ``scripts/platform_urls.py``. Optional: Spalten
``thefork_url``, ``glovo_url``, ``uber_eats_url`` im Blatt ``restaurants``.

Verwendung::

    pip install -r scripts/requirements.txt
    python scripts/excel-to-json.py
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("Fehler: openpyxl nicht installiert.")
    print("Bitte ausfuehren: pip install -r scripts/requirements.txt")
    sys.exit(1)

from platform_urls import (
    enrich_restaurant_links,
    extract_inline_urls_from_row,
    merge_link_lists,
    normalize_platform,
)

# ─── Konfiguration ───────────────────────────────────────────────────────────

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = ROOT / "data-source" / "CeliacSafe_Datenbank_v4.xlsx"
DEFAULT_OUTPUT = ROOT / "src" / "data" / "restaurants.json"
DEFAULT_SHEET = "restaurants"
META_VERSION = "1.0.0"

# Nur verified in die App exportieren (to_be_verified etc. bleiben in Excel/Admin)
INCLUDE_PENDING_FOR_DEV = False

# ─── Felddefinitionen (entspricht src/types/Restaurant.ts) ───────────────────

REQUIRED_FIELDS = (
    "id",
    "name",
    "country_code",
    "region_code",
    "region_name",
    "city",
    "verification_status",
)

STRING_FIELDS = {
    "slug",
    "province",
    "district",
    "postal_code",
    "address_street",
    "venue_type",
    "price_range",
    "national_authority",
    "phone",
    "whatsapp",
    "email",
    "website",
    "menu_url",
    "google_maps_url",
    "apple_maps_url",
    "instagram",
    "facebook",
    "opening_hours",
    "seasonal_closure",
    "description_es",
    "description_en",
    "description_de",
    "featured_image_url",
}

LIST_FIELDS = {
    "cuisine_types",
    "meal_types",
    "verification_methods",
}

BOOLEAN_FIELDS = {
    "face_program",
    "aoecs_certified",
    "is_published",
    "is_hidden",
}

FLOAT_FIELDS = {
    "latitude",
    "longitude",
}

# Excel-Spaltennamen → kanonische JSON-Felder
EXCEL_FIELD_ALIASES: dict[str, list[str]] = {
    "google_maps_url": ["google_maps_deeplink"],
    "apple_maps_url": ["apple_maps_deeplink"],
    "latitude": ["geo_latitude"],
    "longitude": ["geo_longitude"],
}

DATE_FIELDS = {
    "last_verified_at",
}

# Excel-Werte → kanonischer Export-Status in der App
VERIFICATION_STATUS_MAP = {
    "pending_verification": "to_be_verified",
    "to_be_verified": "to_be_verified",
    "in_verification": "in_verification",
    "verified": "verified",
    "verified_100_percent_gluten_free": "verified",
    "verified_100_percent_gluten_free_owner_recheck": "verified",
    "verified_dedicated_gluten_free_recheck_current_wording": "verified",
    "rejected": "rejected",
}


def normalize_header(value: Any) -> str:
    """Wandelt Excel-Spaltenköpfe in snake_case-Feldnamen um."""
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def parse_string(value: Any) -> str | None:
    """Gibt einen bereinigten String zurück oder None bei leeren Werten."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_bool(value: Any) -> bool | None:
    """
    Parst Booleans aus Excel-Zellen.

    Unterstützt echte bool-Werte (TRUE/FALSE), Zahlen (0/1) und Text.
    """
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().upper()
    if text in {"TRUE", "WAHR", "1", "YES", "JA", "SI", "SÍ", "X"}:
        return True
    if text in {"FALSE", "FALSCH", "0", "NO", "NEIN"}:
        return False
    return None


def parse_list(value: Any) -> list[str] | None:
    """
    Parst kommagetrennte Listen (z. B. cuisine_types, meal_types).

    Beispiel: ``"saludable, vegana, internacional"`` → ``["saludable", "vegana", "internacional"]``
    """
    if value is None or value == "":
        return None
    if isinstance(value, list):
        items = [str(item).strip() for item in value if str(item).strip()]
        return items or None
    text = str(value).strip()
    parts = re.split(r"[;,|]", text)
    items = [part.strip() for part in parts if part.strip()]
    return items or None


def parse_float(value: Any) -> float | None:
    """Parst latitude/longitude als Float."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_date(value: Any) -> str | None:
    """
    Parst Datumsfelder als ISO-String ``YYYY-MM-DD``.

    Akzeptiert ``datetime``/``date``-Objekte aus Excel sowie Text im ISO-Format.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return text
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def normalize_price_range(value: Any) -> str | None:
    """Normalisiert Preisklassen auf €-Symbole (€ bis €€€€)."""
    text = parse_string(value)
    if not text:
        return None
    euro_count = text.count("€")
    if euro_count:
        return "€" * min(euro_count, 4)
    return text


def allowed_verification_statuses() -> set[str]:
    """Gibt die erlaubten verification_status-Werte für den Export zurück."""
    statuses = {"verified"}
    if INCLUDE_PENDING_FOR_DEV:
        statuses.update({"pending_verification", "to_be_verified", "in_verification"})
    return statuses


def normalize_verification_status(value: Any) -> str | None:
    """
    Normalisiert den Verifizierungsstatus aus Excel.

    ``pending_verification`` aus der Excel-Datei wird zu ``to_be_verified`` gemappt.
    """
    raw = parse_string(value)
    if not raw:
        return None
    return VERIFICATION_STATUS_MAP.get(raw, raw)


def read_sheet_rows(workbook: Any, sheet_name: str) -> tuple[list[dict[str, Any]], int]:
    """
    Liest ein Excel-Blatt ein.

    Zeile 1 = Spaltennamen, ab Zeile 2 = Datensätze.

    Returns:
        Tuple aus (Datensätze, Anzahl Spalten)
    """
    if sheet_name not in workbook.sheetnames:
        return [], 0

    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = [normalize_header(cell) for cell in next(rows)]
    except StopIteration:
        return [], 0

    column_count = len([header for header in headers if header])
    records: list[dict[str, Any]] = []

    for values in rows:
        if not values or all(value in (None, "") for value in values):
            continue

        record = {
            headers[col_index]: values[col_index]
            for col_index in range(len(headers))
            if headers[col_index]
        }
        records.append(record)

    return records, column_count


def row_to_link(row: dict[str, Any]) -> dict[str, Any] | None:
    """Wandelt eine Zeile aus delivery_links/reservation_links in ein Link-Objekt um."""
    platform = normalize_platform(row.get("platform"))
    if not platform:
        return None

    link: dict[str, Any] = {
        "platform": platform,
        "url": parse_string(row.get("url")) or "",
    }

    is_active = parse_bool(row.get("is_active"))
    if is_active is not None:
        link["is_active"] = is_active

    return link


def group_links_by_restaurant(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """Gruppiert Link-Zeilen nach restaurant_id."""
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in rows:
        restaurant_id = parse_string(row.get("restaurant_id"))
        link = row_to_link(row)
        if restaurant_id and link:
            grouped[restaurant_id].append(link)

    return grouped


def set_if_present(target: dict[str, Any], key: str, value: Any) -> None:
    """Setzt ein Feld nur, wenn der Wert nicht leer ist (kompakteres JSON)."""
    if value is None:
        return
    if isinstance(value, str) and not value:
        return
    if isinstance(value, list) and not value:
        return
    target[key] = value


def apply_excel_field_aliases(row: dict[str, Any]) -> dict[str, Any]:
    """Übernimmt Werte aus alternativen Excel-Spalten (z. B. *_deeplink, geo_*)."""
    merged = dict(row)
    for target, sources in EXCEL_FIELD_ALIASES.items():
        if target in {"latitude", "longitude"}:
            if parse_float(merged.get(target)) is not None:
                continue
        elif parse_string(merged.get(target)):
            continue
        for source in sources:
            if target in {"latitude", "longitude"}:
                value = parse_float(merged.get(source))
                if value is not None:
                    merged[target] = value
                    break
            else:
                value = parse_string(merged.get(source))
                if value:
                    merged[target] = value
                    break
    return merged


def build_geocoding_lookup(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Indexiert Zeilen aus geocoding_all_129 nach restaurant id."""
    lookup: dict[str, dict[str, Any]] = {}
    for row in rows:
        restaurant_id = parse_string(row.get("id"))
        if restaurant_id:
            lookup[restaurant_id] = row
    return lookup


def apply_geocoding_lookup(row: dict[str, Any], lookup: dict[str, dict[str, Any]]) -> dict[str, Any]:
    """Ergänzt Koordinaten und Maps-Links aus dem Geocoding-Blatt."""
    merged = apply_excel_field_aliases(row)
    restaurant_id = parse_string(merged.get("id"))
    if not restaurant_id:
        return merged

    geo_row = lookup.get(restaurant_id)
    if not geo_row:
        return merged

    merged = apply_excel_field_aliases({**geo_row, **merged})

    for target, sources in EXCEL_FIELD_ALIASES.items():
        if target in {"latitude", "longitude"}:
            if parse_float(merged.get(target)) is not None:
                continue
            for source in sources:
                value = parse_float(geo_row.get(source))
                if value is not None:
                    merged[target] = value
                    break
        elif not parse_string(merged.get(target)):
            for source in sources:
                value = parse_string(geo_row.get(source))
                if value:
                    merged[target] = value
                    break

    return merged


def row_to_restaurant(row: dict[str, Any], row_number: int) -> dict[str, Any]:
    """
    Wandelt eine Excel-Zeile in ein Restaurant-Objekt um.

    Pflichtfelder werden immer gesetzt; leere optionale Felder werden weggelassen.
    """
    row = apply_excel_field_aliases(row)
    restaurant: dict[str, Any] = {}

    for field in REQUIRED_FIELDS:
        if field == "verification_status":
            raw_status = parse_string(row.get(field))
            if raw_status is None:
                raise ValueError(
                    f"Zeile {row_number}: Pflichtfeld '{field}' fehlt oder ist leer."
                )
            restaurant[field] = normalize_verification_status(raw_status) or raw_status
            continue

        value = parse_string(row.get(field))
        if value is None:
            raise ValueError(
                f"Zeile {row_number}: Pflichtfeld '{field}' fehlt oder ist leer."
            )
        restaurant[field] = value

    for field in STRING_FIELDS:
        if field == "price_range":
            set_if_present(restaurant, field, normalize_price_range(row.get(field)))
        else:
            set_if_present(restaurant, field, parse_string(row.get(field)))

    for field in LIST_FIELDS:
        set_if_present(restaurant, field, parse_list(row.get(field)))

    for field in BOOLEAN_FIELDS:
        parsed = parse_bool(row.get(field))
        if parsed is not None:
            restaurant[field] = parsed

    for field in FLOAT_FIELDS:
        set_if_present(restaurant, field, parse_float(row.get(field)))

    for field in DATE_FIELDS:
        set_if_present(restaurant, field, parse_date(row.get(field)))

    return restaurant


def convert_workbook(path: Path, sheet_name: str) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """
    Konvertiert die Excel-Datei in eine gefilterte Restaurant-Liste.

    Returns:
        Tuple aus (restaurants, statistik)
    """
    workbook = load_workbook(path, read_only=True, data_only=True)

    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        workbook.close()
        raise ValueError(
            f"Blatt '{sheet_name}' nicht gefunden. Verfuegbar: {available}"
        )

    restaurant_rows, column_count = read_sheet_rows(workbook, sheet_name)
    geocoding_lookup = build_geocoding_lookup(
        read_sheet_rows(workbook, "geocoding_all_129")[0]
    )
    delivery_by_id = group_links_by_restaurant(read_sheet_rows(workbook, "delivery_links")[0])
    reservation_by_id = group_links_by_restaurant(
        read_sheet_rows(workbook, "reservation_links")[0]
    )
    workbook.close()

    allowed_statuses = allowed_verification_statuses()
    restaurants: list[dict[str, Any]] = []
    stats = {
        "columns": column_count,
        "read": len(restaurant_rows),
        "delivery_restaurants": len(delivery_by_id),
        "reservation_restaurants": len(reservation_by_id),
        "delivery_urls_enriched": 0,
        "reservation_urls_enriched": 0,
        "exported": 0,
        "skipped_status": 0,
        "skipped_invalid": 0,
        "with_coordinates": 0,
    }

    for index, row in enumerate(restaurant_rows, start=2):
        row = apply_geocoding_lookup(row, geocoding_lookup)
        status = normalize_verification_status(row.get("verification_status"))
        if not status or status not in allowed_statuses:
            stats["skipped_status"] += 1
            continue

        try:
            restaurant = row_to_restaurant(row, index)
        except ValueError:
            stats["skipped_invalid"] += 1
            continue

        if restaurant.get("latitude") is not None and restaurant.get("longitude") is not None:
            stats["with_coordinates"] += 1

        restaurant_id = restaurant["id"]

        inline = extract_inline_urls_from_row(row)
        delivery_links = merge_link_lists(
            delivery_by_id.get(restaurant_id, []),
            inline["delivery"],
        )
        reservation_links = merge_link_lists(
            reservation_by_id.get(restaurant_id, []),
            inline["reservation"],
        )

        delivery_links, reservation_links, enrich_stats = enrich_restaurant_links(
            restaurant,
            delivery_links or None,
            reservation_links or None,
        )
        stats["delivery_urls_enriched"] += enrich_stats["delivery_enriched"]
        stats["reservation_urls_enriched"] += enrich_stats["reservation_enriched"]

        if delivery_links:
            restaurant["delivery_links"] = delivery_links

        if reservation_links:
            restaurant["reservation_links"] = reservation_links

        restaurants.append(restaurant)
        stats["exported"] += 1

    return restaurants, stats


def build_output(restaurants: list[dict[str, Any]], source: str) -> dict[str, Any]:
    """Erstellt die finale JSON-Struktur mit ``_meta``-Block und ``restaurants``-Array."""
    return {
        "_meta": {
            "version": META_VERSION,
            "count": len(restaurants),
            "source": source,
        },
        "restaurants": restaurants,
    }


def print_status(
    stats: dict[str, int],
    input_path: Path,
    output_path: Path,
) -> None:
    """Gibt eine deutschsprachige Statusausgabe auf der Konsole aus."""
    print(f"Lade Excel-Datei: {input_path}")
    print(f"Spalten gefunden: {stats['columns']}")
    print(f"Datenzeilen: {stats['read']}")
    print(f"Delivery-Links fuer {stats['delivery_restaurants']} Restaurants")
    print(f"Reservierungs-Links fuer {stats['reservation_restaurants']} Restaurants")
    print(
        f"Automatisch ergaenzte URLs: "
        f"{stats.get('delivery_urls_enriched', 0)} Lieferung, "
        f"{stats.get('reservation_urls_enriched', 0)} Reservierung"
    )
    print()
    print("=== ERGEBNIS ===")
    print(f"Exportiert: {stats['exported']} Restaurants")
    print(f"Uebersprungen wegen Status: {stats['skipped_status']}")
    print(f"Uebersprungen wegen ungueltiger Daten: {stats['skipped_invalid']}")
    print(f"Mit Koordinaten: {stats.get('with_coordinates', 0)}")
    print(f"Ausgabedatei: {output_path}")

    if output_path.exists():
        size_kb = output_path.stat().st_size / 1024
        print(f"Dateigroesse: {size_kb:.1f} KB")

    if INCLUDE_PENDING_FOR_DEV:
        print("Hinweis: INCLUDE_PENDING_FOR_DEV=True (Entwicklungsmodus aktiv)")

    if stats["exported"] == 0:
        print()
        print("Warnung: Keine Datensaetze exportiert.")
        print("Tipp: Setze INCLUDE_PENDING_FOR_DEV=True fuer die Entwicklung.")


def merge_workbooks(input_paths: list[Path], sheet_name: str) -> tuple[list[dict[str, Any]], dict[str, int], str]:
    """Liest mehrere Excel-Dateien und vereinigt Restaurants (nach id, spätere Datei gewinnt)."""
    by_id: dict[str, dict[str, Any]] = {}
    combined_stats = {
        "columns": 0,
        "read": 0,
        "delivery_restaurants": 0,
        "reservation_restaurants": 0,
        "delivery_urls_enriched": 0,
        "reservation_urls_enriched": 0,
        "exported": 0,
        "skipped_status": 0,
        "skipped_invalid": 0,
        "with_coordinates": 0,
    }
    source_names: list[str] = []

    for input_path in input_paths:
        restaurants, stats = convert_workbook(input_path, sheet_name)
        source_names.append(input_path.name)
        combined_stats["columns"] = max(combined_stats["columns"], stats["columns"])
        for key in (
            "read",
            "delivery_restaurants",
            "reservation_restaurants",
            "delivery_urls_enriched",
            "reservation_urls_enriched",
            "skipped_status",
            "skipped_invalid",
            "with_coordinates",
        ):
            combined_stats[key] += stats[key]
        for restaurant in restaurants:
            by_id[restaurant["id"]] = restaurant

    combined_stats["exported"] = len(by_id)
    source = "data-source/" + " + ".join(source_names)
    return list(by_id.values()), combined_stats, source


def main() -> None:
    """Einstiegspunkt: Excel einlesen, filtern, JSON schreiben."""
    parser = argparse.ArgumentParser(
        description="Exportiert CeliacSafe-Restaurants aus Excel nach JSON.",
        epilog=(
            "Beispiele:\n"
            "  python scripts/excel-to-json.py\n"
            "  python scripts/excel-to-json.py data-source/CeliacSafe_Datenbank_v3.xlsx\n"
            "  python scripts/excel-to-json.py data-source/CeliacSafe_Datenbank_v3.xlsx src/data/restaurants.json"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "input_pos",
        nargs="?",
        type=Path,
        default=None,
        help="Excel-Eingabedatei (optional, Positionsargument)",
    )
    parser.add_argument(
        "output_pos",
        nargs="?",
        type=Path,
        default=None,
        help="JSON-Ausgabedatei (optional, Positionsargument)",
    )
    parser.add_argument(
        "--input",
        "-i",
        action="append",
        type=Path,
        default=None,
        help=f"Excel-Eingabedatei (mehrfach für Merge; Standard: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=None,
        help=f"JSON-Ausgabedatei (Standard: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help=f"Name des Restaurant-Blattes (Standard: {DEFAULT_SHEET})",
    )
    args = parser.parse_args()

    output_path = args.output or args.output_pos or DEFAULT_OUTPUT

    if args.input:
        input_paths = args.input
    elif args.input_pos:
        input_paths = [args.input_pos]
    else:
        input_paths = [DEFAULT_INPUT]

    for input_path in input_paths:
        if not input_path.exists():
            print(f"Fehler: Excel-Datei nicht gefunden: {input_path}")
            sys.exit(1)

    if len(input_paths) == 1:
        restaurants, stats = convert_workbook(input_paths[0], args.sheet)
        source = f"data-source/{input_paths[0].name}"
    else:
        restaurants, stats, source = merge_workbooks(input_paths, args.sheet)

    payload = build_output(restaurants, source)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print_status(stats, input_paths[0] if len(input_paths) == 1 else Path("merge"), output_path)
    if len(input_paths) > 1:
        print(f"Quellen: {', '.join(p.name for p in input_paths)}")


if __name__ == "__main__":
    main()
