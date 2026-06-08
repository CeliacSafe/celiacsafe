#!/usr/bin/env python3
"""
CeliacSafe — Geocoding fuer Restaurant-Koordinaten in der Excel-Datenbank.

Liest eine Excel-Datei (z. B. v3), ergaenzt fehlende latitude/longitude per
Nominatim (OpenStreetMap) und schreibt eine neue Datei (z. B. v4).

Verwendung::

    cd scripts
    python geocode.py ../data-source/CeliacSafe_Datenbank_v3.xlsx ../data-source/CeliacSafe_Datenbank_v4.xlsx

Bei Verbindungsproblemen oder 403 von Nominatim::

    python geocode.py ../data-source/CeliacSafe_Datenbank_v3.xlsx ../data-source/CeliacSafe_Datenbank_v4.xlsx --fallback-only
"""

from __future__ import annotations

import argparse
import json
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("Fehler: openpyxl nicht installiert.")
    print("Bitte ausfuehren: pip install -r scripts/requirements.txt")
    sys.exit(1)

NOMINATIM_DELAY_SEC = 1.2
USER_AGENT = "CeliacSafe/1.0 (educational; geocoding for gluten-free restaurant map)"
DEFAULT_SHEET = "restaurants"


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def has_coordinates(lat: Any, lng: Any) -> bool:
    lat_f = parse_float(lat)
    lng_f = parse_float(lng)
    return lat_f is not None and lng_f is not None


def nominatim_search(query: str) -> tuple[float, float] | None:
    """Fragt Nominatim nach einer Adresse ab (max. 1 Treffer, nur Spanien)."""
    params = urllib.parse.urlencode(
        {
            "q": query,
            "format": "json",
            "limit": 1,
            "countrycodes": "es",
        }
    )
    url = f"https://nominatim.openstreetmap.org/search?{params}"
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})

    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError) as exc:
        print(f"  Nominatim-Fehler fuer '{query}': {exc}")
        return None

    if not payload:
        return None

    try:
        return float(payload[0]["lat"]), float(payload[0]["lon"])
    except (KeyError, TypeError, ValueError):
        return None


def parse_string(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def resolve_coordinate_columns(headers: list[str]) -> tuple[str, str]:
    if "latitude" in headers and "longitude" in headers:
        return "latitude", "longitude"
    if "geo_latitude" in headers and "geo_longitude" in headers:
        return "geo_latitude", "geo_longitude"
    raise ValueError(
        "Spalten 'latitude'/'longitude' oder 'geo_latitude'/'geo_longitude' fehlen im restaurants-Blatt."
    )


def build_search_query(row: dict[str, Any]) -> str:
    query = parse_string(row.get("geocoding_query"))
    if query:
        return query
    return build_address_query(row)


def build_address_query(row: dict[str, Any]) -> str:
    parts: list[str] = []
    street = row.get("address_street")
    postal = row.get("postal_code")
    city = row.get("city")
    province = row.get("province")

    if street:
        parts.append(str(street).strip())
    if postal:
        parts.append(str(postal).strip())
    if city:
        parts.append(str(city).strip())
    if province and province != city:
        parts.append(str(province).strip())
    parts.append("Espana")
    return ", ".join(parts)


def build_city_query(city: str) -> str:
    return f"{city.strip()}, Espana"


def geocode_workbook(
    input_path: Path,
    output_path: Path,
    *,
    sheet_name: str,
    fallback_only: bool,
) -> dict[str, int]:
    workbook = load_workbook(input_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Blatt '{sheet_name}' nicht gefunden.")

    worksheet = workbook[sheet_name]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(cell) for cell in header_row]

    lat_key, lng_key = resolve_coordinate_columns(headers)
    lat_col = headers.index(lat_key) + 1
    lng_col = headers.index(lng_key) + 1
    geo_sheet = (
        workbook["geocoding_all_129"]
        if "geocoding_all_129" in workbook.sheetnames
        else None
    )
    geo_headers: list[str] = []
    geo_lat_col = geo_lng_col = None
    if geo_sheet is not None:
        geo_header_row = next(geo_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        geo_headers = [normalize_header(cell) for cell in geo_header_row]
        if "geo_latitude" in geo_headers and "geo_longitude" in geo_headers:
            geo_lat_col = geo_headers.index("geo_latitude") + 1
            geo_lng_col = geo_headers.index("geo_longitude") + 1

    stats = {
        "already_had_coords": 0,
        "online_geocoding": 0,
        "city_fallback": 0,
        "failed": 0,
    }
    city_cache: dict[str, tuple[float, float] | None] = {}

    for row_index, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not values or all(value in (None, "") for value in values):
            continue

        row = {
            headers[col_index]: values[col_index]
            for col_index in range(len(headers))
            if headers[col_index]
        }
        name = row.get("name") or row.get("id") or f"Zeile {row_index}"
        lat_cell = worksheet.cell(row=row_index, column=lat_col)
        lng_cell = worksheet.cell(row=row_index, column=lng_col)

        if has_coordinates(lat_cell.value, lng_cell.value):
            stats["already_had_coords"] += 1
            continue

        coords: tuple[float, float] | None = None

        if not fallback_only:
            query = build_search_query(row)
            print(f"[{row_index}] Online: {name} — {query}")
            coords = nominatim_search(query)
            time.sleep(NOMINATIM_DELAY_SEC)
            if coords:
                stats["online_geocoding"] += 1

        if coords is None:
            city = row.get("city")
            if city:
                city_key = str(city).strip().lower()
                if city_key not in city_cache:
                    city_query = build_city_query(str(city))
                    print(f"[{row_index}] Stadt-Fallback: {name} — {city_query}")
                    city_cache[city_key] = nominatim_search(city_query)
                    time.sleep(NOMINATIM_DELAY_SEC)
                coords = city_cache[city_key]
                if coords and fallback_only:
                    stats["city_fallback"] += 1
                elif coords and not fallback_only:
                    stats["city_fallback"] += 1

        if coords:
            lat_cell.value = round(coords[0], 6)
            lng_cell.value = round(coords[1], 6)
            if geo_sheet is not None and geo_lat_col and geo_lng_col:
                restaurant_id = parse_string(row.get("id"))
                if restaurant_id:
                    for geo_row_index, geo_values in enumerate(
                        geo_sheet.iter_rows(min_row=2, values_only=True),
                        start=2,
                    ):
                        geo_row = {
                            geo_headers[col_index]: geo_values[col_index]
                            for col_index in range(len(geo_headers))
                            if geo_headers[col_index]
                        }
                        if parse_string(geo_row.get("id")) != restaurant_id:
                            continue
                        geo_sheet.cell(row=geo_row_index, column=geo_lat_col).value = lat_cell.value
                        geo_sheet.cell(row=geo_row_index, column=geo_lng_col).value = lng_cell.value
                        break
        else:
            stats["failed"] += 1
            print(f"[{row_index}] FEHLGESCHLAGEN: {name}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Ergaenzt latitude/longitude in der CeliacSafe-Excel-Datei.",
    )
    parser.add_argument("input", type=Path, help="Eingabe-xlsx (z. B. v3)")
    parser.add_argument("output", type=Path, help="Ausgabe-xlsx (z. B. v4)")
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help=f"Blattname (Standard: {DEFAULT_SHEET})",
    )
    parser.add_argument(
        "--fallback-only",
        action="store_true",
        help="Nur Stadt-Mittelpunkte per Nominatim, keine Adress-Suche",
    )
    args = parser.parse_args()

    if not args.input.exists():
        print(f"Fehler: Datei nicht gefunden: {args.input}")
        sys.exit(1)

    mode = "Stadt-Fallback only" if args.fallback_only else "Online + Stadt-Fallback"
    print(f"Geocoding: {args.input.name} -> {args.output.name} ({mode})")
    print()

    stats = geocode_workbook(
        args.input,
        args.output,
        sheet_name=args.sheet,
        fallback_only=args.fallback_only,
    )

    print()
    print("=== ERGEBNIS ===")
    print(f"Bereits mit Koordinaten: {stats['already_had_coords']}")
    print(f"Online-Geocoding: {stats['online_geocoding']}")
    print(f"Stadt-Fallback: {stats['city_fallback']}")
    print(f"Fehlgeschlagen: {stats['failed']}")
    print(f"Ausgabedatei: {args.output}")

    if stats["failed"] > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
