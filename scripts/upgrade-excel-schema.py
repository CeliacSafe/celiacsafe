#!/usr/bin/env python3
"""
Aktualisiert eine bestehende CeliacSafe-Excel-Datei auf das aktuelle Spaltenschema
(menu_url, google_maps_url, apple_maps_url, is_published, is_hidden, …).

Standard: liest v4_enriched (oder v3), schreibt data-source/CeliacSafe_Datenbank_v4.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = ROOT / "data-source" / "CeliacSafe_Datenbank_v4_enriched.xlsx"
FALLBACK_INPUT = ROOT / "data-source" / "CeliacSafe_Datenbank_v3.xlsx"
DEFAULT_OUTPUT = ROOT / "data-source" / "CeliacSafe_Datenbank_v4.xlsx"

RESTAURANT_HEADERS = [
    "id",
    "name",
    "slug",
    "country_code",
    "region_code",
    "region_name",
    "province",
    "city",
    "district",
    "postal_code",
    "address_street",
    "latitude",
    "longitude",
    "google_maps_url",
    "apple_maps_url",
    "venue_type",
    "cuisine_types",
    "price_range",
    "meal_types",
    "verification_status",
    "verification_methods",
    "last_verified_at",
    "face_program",
    "aoecs_certified",
    "national_authority",
    "phone",
    "whatsapp",
    "email",
    "website",
    "menu_url",
    "instagram",
    "facebook",
    "opening_hours",
    "seasonal_closure",
    "description_de",
    "description_en",
    "description_es",
    "featured_image_url",
    "is_published",
    "is_hidden",
    "thefork_url",
    "glovo_url",
    "uber_eats_url",
    "just_eat_url",
]

# Zusätzliche Spalten aus älteren Excel-Versionen (am Ende beibehalten)
LEGACY_TAIL_HEADERS = [
    "verified_by",
    "verification_notes",
    "rejection_reason",
    "data_source_type",
    "data_source_notes",
    "created_at",
    "updated_at",
]

HEADER_FONT = Font(bold=True)


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def read_sheet_dicts(ws: Worksheet) -> tuple[list[str], list[dict[str, Any]]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [normalize_header(c) for c in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        record: dict[str, Any] = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            record[key] = row[idx] if idx < len(row) else None
        if record.get("id") or record.get("name"):
            records.append(record)
    return headers, records


def build_target_headers(existing: list[str]) -> list[str]:
    seen: set[str] = set()
    target: list[str] = []
    for h in RESTAURANT_HEADERS + LEGACY_TAIL_HEADERS:
        if h and h not in seen:
            target.append(h)
            seen.add(h)
    for h in existing:
        if h and h not in seen:
            target.append(h)
            seen.add(h)
    return target


def write_restaurants_sheet(ws: Worksheet, records: list[dict[str, Any]], headers: list[str]) -> None:
    ws.delete_rows(1, ws.max_row or 1)
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(28, len(name) + 2))
    for row_idx, record in enumerate(records, start=2):
        for col, name in enumerate(headers, start=1):
            value = record.get(name)
            if value is not None and value != "":
                ws.cell(row=row_idx, column=col, value=value)


def upgrade_workbook(input_path: Path, output_path: Path) -> None:
    wb = load_workbook(input_path)
    if "restaurants" not in wb.sheetnames:
        raise ValueError(f"Blatt 'restaurants' fehlt in {input_path.name}")

    old_headers, records = read_sheet_dicts(wb["restaurants"])
    target_headers = build_target_headers(old_headers)

    for record in records:
        if record.get("is_published") in (None, ""):
            record["is_published"] = True
        if record.get("is_hidden") in (None, ""):
            record["is_hidden"] = False
        for key in ("google_maps_url", "apple_maps_url", "menu_url"):
            record.setdefault(key, None)

    write_restaurants_sheet(wb["restaurants"], records, target_headers)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Aktualisiert: {output_path}")
    print(f"  restaurants: {len(records)} Zeilen, {len(target_headers)} Spalten")
    print(f"  Neu: google_maps_url, apple_maps_url, menu_url (leer — bitte pflegen)")


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel-Schema auf aktuelle Spalten bringen")
    parser.add_argument("-i", "--input", type=Path, default=None)
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    input_path = args.input
    if input_path is None:
        input_path = DEFAULT_INPUT if DEFAULT_INPUT.exists() else FALLBACK_INPUT
    if not input_path.exists():
        raise SystemExit(f"Eingabedatei nicht gefunden: {input_path}")

    upgrade_workbook(input_path, args.output)


if __name__ == "__main__":
    main()
