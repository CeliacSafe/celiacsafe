#!/usr/bin/env python3
"""
CeliacSafe — Beschreibungen vereinheitlichen (DE/EN/ES).

1. Ersetzt Audit-/App-Tauglichkeitstexte durch speisenorientierte Kurzbeschreibungen.
2. Leitet aus den deutschen Template-Texten konsistente EN- und ES-Beschreibungen ab.

Die Texte beschreiben Angebot und Küche — nicht Verifizierung oder App-Eignung.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("Fehler: openpyxl nicht installiert.")
    print("Bitte: pip install -r scripts/requirements.txt")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_TARGETS = [
    ROOT / "data-source" / "CeliacSafe_Datenbank_v4_Spain_129_geocoded.xlsx",
    ROOT / "data-source" / "CeliacSafe_Datenbank_v4_Germany_Delta_Consolidated_geocoded.xlsx",
]

# ─── Kategorie-Templates (Speisenangebot, keine Verifizierungs-Metadaten) ───

CATEGORY_TEMPLATES: dict[str, dict[str, str]] = {
    "tapas_restaurant": {
        "de": "Restaurant in {city} mit Tapas, mediterranen Gerichten, Fisch-/Fleischspeisen und spanisch inspirierten Klassikern.",
        "en": "Restaurant in {city} serving tapas, Mediterranean dishes, fish and meat specialities, and Spanish-inspired classics.",
        "es": "Restaurante en {city} con tapas, platos mediterráneos, pescados y carnes, y clásicos de inspiración española.",
    },
    "healthy": {
        "de": "Healthy-Restaurant/Café in {city} mit Bowls, Brunchgerichten, Salaten, herzhaften Speisen, Kuchen und Getränken.",
        "en": "Health-focused restaurant and café in {city} offering bowls, brunch dishes, salads, savoury plates, cakes, and drinks.",
        "es": "Restaurante y cafetería saludable en {city} con bowls, platos de brunch, ensaladas, platos salados, tartas y bebidas.",
    },
    "italian": {
        "de": "Italienisches Restaurant in {city} mit Pizza, Pasta, Antipasti und weiteren mediterranen Gerichten.",
        "en": "Italian restaurant in {city} serving pizza, pasta, antipasti, and other Mediterranean dishes.",
        "es": "Restaurante italiano en {city} con pizza, pasta, antipasti y otros platos mediterráneos.",
    },
    "bakery": {
        "de": "Bäckerei/Patisserie in {city} mit Brot, Brötchen, Kuchen, Torten, Gebäck, süßen Backwaren und teilweise herzhaften Snacks.",
        "en": "Bakery and patisserie in {city} offering bread, rolls, cakes, tarts, pastries, sweet baked goods, and a selection of savoury snacks.",
        "es": "Panadería y pastelería en {city} con pan, panecillos, tartas, pasteles, bollería, dulces y algunos snacks salados.",
    },
    "bistro": {
        "de": "Restaurant/Bistro in {city} mit Vorspeisen, Hauptgerichten, Desserts und je nach Konzept regionaler, mediterraner oder internationaler Küche.",
        "en": "Restaurant and bistro in {city} serving starters, main courses, and desserts, with regional, Mediterranean, or international cuisine depending on the concept.",
        "es": "Restaurante y bistró en {city} con entrantes, platos principales y postres; cocina regional, mediterránea o internacional según el concepto.",
    },
    "takeaway": {
        "de": "Take-away-Konzept in {city} mit vorbereiteten Speisen, Snacks, Backwaren und Gerichten zum Mitnehmen.",
        "en": "Takeaway concept in {city} offering prepared meals, snacks, baked goods, and dishes to go.",
        "es": "Concepto take-away en {city} con comidas preparadas, snacks, productos de panadería y platos para llevar.",
    },
    "vegan": {
        "de": "Veganes bzw. plant-based Café/Restaurant in {city} mit Bowls, Kuchen, Frühstücksgerichten, Snacks und kreativer pflanzlicher Küche.",
        "en": "Vegan, plant-based café and restaurant in {city} offering bowls, cakes, breakfast dishes, snacks, and creative plant-based cuisine.",
        "es": "Cafetería y restaurante vegano en {city} con bowls, tartas, desayunos, snacks y cocina vegetal creativa.",
    },
    "japanese": {
        "de": "Japanisches Restaurant in {city} mit Sushi, warmen japanischen Gerichten, Reisgerichten und Desserts.",
        "en": "Japanese restaurant in {city} serving sushi, hot Japanese dishes, rice dishes, and desserts.",
        "es": "Restaurante japonés en {city} con sushi, platos japoneses calientes, platos de arroz y postres.",
    },
    "croquet": {
        "de": "Croquetería in {city} mit verschiedenen Croquetas, kleinen herzhaften Gerichten und Take-away-Angebot.",
        "en": "Croquette specialist in {city} offering a variety of croquetas, small savoury dishes, and takeaway.",
        "es": "Croquetería en {city} con croquetas variadas, pequeños platos salados y opción para llevar.",
    },
    "burger": {
        "de": "Burgerrestaurant in {city} mit Burgern, Beilagen, Saucen und Casual-Food-Gerichten.",
        "en": "Burger restaurant in {city} serving burgers, sides, sauces, and casual food.",
        "es": "Hamburguesería en {city} con hamburguesas, acompañamientos, salsas y comida informal.",
    },
    "mexican": {
        "de": "Mexikanisches Restaurant in {city} mit Tacos, Nachos, Salsas, Bowls und weiteren mexikanischen Klassikern.",
        "en": "Mexican restaurant in {city} serving tacos, nachos, salsas, bowls, and other Mexican classics.",
        "es": "Restaurante mexicano en {city} con tacos, nachos, salsas, bowls y otros clásicos mexicanos.",
    },
    "cafe": {
        "de": "Café in {city} mit Kaffee, Frühstück, Kuchen, Gebäck, Brunchgerichten und kleinen Speisen.",
        "en": "Café in {city} offering coffee, breakfast, cakes, pastries, brunch dishes, and light bites.",
        "es": "Cafetería en {city} con café, desayunos, tartas, bollería, platos de brunch y comidas ligeras.",
    },
    "pizzeria": {
        "de": "Pizzeria in {city} mit Pizza, herzhaften Ofengerichten und italienisch inspirierten Speisen.",
        "en": "Pizzeria in {city} serving pizza, savoury oven-baked dishes, and Italian-inspired food.",
        "es": "Pizzería en {city} con pizza, platos al horno y cocina de inspiración italiana.",
    },
    "regional": {
        "de": "Restaurant in {city} mit regionaler Küche, Hauptgerichten, Vorspeisen, Desserts und saisonalen Speisen.",
        "en": "Restaurant in {city} serving regional cuisine, main courses, starters, desserts, and seasonal dishes.",
        "es": "Restaurante en {city} con cocina regional, platos principales, entrantes, postres y platos de temporada.",
    },
    "latin": {
        "de": "Lateinamerikanisches Restaurant/Café in {city} mit Tapioca, Arepas, Empanadas, Bowls und herzhaften Snacks.",
        "en": "Latin American restaurant and café in {city} serving tapioca, arepas, empanadas, bowls, and savoury snacks.",
        "es": "Restaurante y café latinoamericano en {city} con tapioca, arepas, empanadas, bowls y snacks salados.",
    },
}

TEMPLATES: list[dict[str, str]] = [
    {"key": key, "de": value["de"], "en": value["en"], "es": value["es"]}
    for key, value in CATEGORY_TEMPLATES.items()
]

# Marker fuer interne Audit-/App-Tauglichkeitstexte (keine Gastro-Beschreibung)
AUDIT_MARKERS: tuple[re.Pattern[str], ...] = tuple(
    re.compile(pattern, re.IGNORECASE)
    for pattern in [
        r"offiziell(er|en)?\s+auftritt",
        r"als\s+100\s*%\s*gluten",
        r"100\s*%\s*gluten\s*frei",
        r"100\s*%-?\s*gluten",
        r"100\s*%\s*sin\s+gluten",
        r"gluten\s*free",
        r"für die app",
        r"fuer die app",
        r"app\s*geeignet",
        r"geeignet\s+für",
        r"recommended_app",
        r"verifizier",
        r"verification",
        r"kontamin",
        r"cross.?contamin",
        r"beschrieben",
        r"weist\s+.*\s+aus",
        r"führt\s+100",
        r"listung",
        r"listing",
        r"celiac\s*safe",
        r"zöliak",
        r"coeliac",
        r"face.?program",
        r"aoecs",
        r"laktosefrei\s+beschrieben",
        r"sin\s+lactosa\s+beschrieben",
    ]
)

# ─── Gängige Exonyme für Städtenamen in EN/ES ───────────────────────────────

CITY_EXONYMS_EN: dict[str, str] = {
    "München": "Munich",
    "Köln": "Cologne",
    "Sevilla": "Seville",
}

CITY_EXONYMS_ES: dict[str, str] = {
    "Berlin": "Berlín",
    "München": "Múnich",
    "Köln": "Colonia",
    "Hamburg": "Hamburgo",
    "Frankfurt am Main": "Fráncfort del Meno",
    "Aachen": "Aquisgrán",
    "Regensburg": "Ratisbona",
}

COMPILED_TEMPLATES = [
    {
        "pattern": re.compile(
            "^"
            + re.escape(template["de"].format(city="__CITY__"))
            .replace("__CITY__", "(?P<city>.+?)")
            + "$"
        ),
        "en": template["en"],
        "es": template["es"],
    }
    for template in TEMPLATES
]

# ─── DE-Seed-Restaurants: Audit-Text → richtiges Kategorie-Template ─────────

DE_TEMPLATE_BAKERY = CATEGORY_TEMPLATES["bakery"]["de"]
DE_TEMPLATE_VEGAN = CATEGORY_TEMPLATES["vegan"]["de"]

SEED_DE_FIXES: dict[str, str] = {
    "DE-011": DE_TEMPLATE_BAKERY.format(city="Berlin"),
    "DE-012": DE_TEMPLATE_BAKERY.format(city="Berlin"),
    "DE-013": DE_TEMPLATE_BAKERY.format(city="Berlin"),
    "DE-014": DE_TEMPLATE_BAKERY.format(city="Berlin"),
    "DE-016": DE_TEMPLATE_BAKERY.format(city="Essen"),
    "DE-017": DE_TEMPLATE_BAKERY.format(city="Düsseldorf"),
    "DE-022": DE_TEMPLATE_VEGAN.format(city="Frankfurt am Main"),
    "DE-023": DE_TEMPLATE_BAKERY.format(city="Wiesbaden"),
    "DE-025": DE_TEMPLATE_BAKERY.format(city="Lorch"),
}


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def is_audit_style_description(text: str) -> bool:
    """True bei internen Audit-/Verifizierungstexten statt Gastro-Beschreibung."""
    cleaned = text.strip()
    if not cleaned:
        return False
    if any(marker.search(cleaned) for marker in AUDIT_MARKERS):
        return True
    # Kurze Meta-Sätze mit Semikolon (typisch Audit-Notizen)
    if ";" in cleaned and len(cleaned) < 220:
        tail = cleaned.split(";", 1)[1].lower()
        if any(word in tail for word in ("gluten", "laktose", "verif", "auftritt", "beschrieb")):
            return True
    return False


def classify_venue_category(venue_type: Any, name: Any) -> str:
    venue = str(venue_type or "").lower()
    title = str(name or "").lower()

    if any(token in venue for token in ("pizza", "pizzeria")) or "pizza" in title:
        return "pizzeria"
    if "burger" in venue or "burger" in title:
        return "burger"
    if "croquet" in venue or "croquet" in title:
        return "croquet"
    if any(token in venue for token in ("japan", "sushi")) or "sushi" in title:
        return "japanese"
    if "mexic" in venue or "taco" in title:
        return "mexican"
    if any(token in venue for token in ("vegan", "plant", "raw cake", "raw cakes")):
        return "vegan"
    if any(token in venue for token in ("healthy", "superfood", "bowl")):
        return "healthy"
    if "italian" in venue or "vietnamese" in venue:
        return "italian" if "italian" in venue else "bistro"
    if any(token in venue for token in ("latin", "brazilian", "tapioca", "arepa")):
        return "latin"
    if any(
        token in venue
        for token in ("bakery", "bäckerei", "backerei", "pastry", "patisserie", "obrador", "churrer")
    ) or any(token in title for token in ("bakery", "bäckerei", "backerei", "patisserie")):
        return "bakery"
    if any(token in venue for token in ("café", "cafe", "coffee")) or "coffee" in title:
        return "cafe"
    if any(token in venue for token in ("takeaway", "take-away", "to-go", "to go")):
        return "takeaway"
    if "tapas" in title or "sidrer" in venue:
        return "tapas_restaurant"
    if "restaurant" in venue:
        return "bistro"
    return "bistro"


def build_category_copy(category: str, city: str) -> dict[str, str]:
    template = CATEGORY_TEMPLATES.get(category, CATEGORY_TEMPLATES["bistro"])
    city_en = CITY_EXONYMS_EN.get(city, city)
    city_es = CITY_EXONYMS_ES.get(city, city)
    return {
        "de": template["de"].format(city=city),
        "en": template["en"].format(city=city_en),
        "es": template["es"].format(city=city_es),
    }


def krumcoffee_copy(city: str) -> dict[str, str]:
    city_en = CITY_EXONYMS_EN.get(city, city)
    city_es = CITY_EXONYMS_ES.get(city, city)
    return {
        "de": (
            f"KrümCoffee-Café in {city} mit Kaffee, Frühstück, Brunch, Kuchen, "
            f"Gebäck und herzhaften Snacks."
        ),
        "en": (
            f"KrümCoffee café in {city_en} serving coffee, breakfast, brunch, "
            f"cakes, pastries, and savoury snacks."
        ),
        "es": (
            f"Cafetería KrümCoffee en {city_es} con café, desayuno, brunch, "
            f"tartas, bollería y snacks salados."
        ),
    }


def is_krumcoffee(name: Any) -> bool:
    text = str(name or "").lower()
    return "krümcoffee" in text or "krumcoffee" in text or "krüm coffee" in text


def translate_from_german(description_de: str) -> tuple[str, str] | None:
    """Gibt (en, es) zurueck, wenn der DE-Text einem Template entspricht."""
    text = description_de.strip()
    for template in COMPILED_TEMPLATES:
        match = template["pattern"].match(text)
        if match:
            city = match.group("city")
            return (
                template["en"].format(city=CITY_EXONYMS_EN.get(city, city)),
                template["es"].format(city=CITY_EXONYMS_ES.get(city, city)),
            )
    return None


def apply_to_workbook(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path)
    stats = {
        "seed_de_fixed": 0,
        "audit_replaced": 0,
        "special_copy": 0,
        "en_set": 0,
        "es_set": 0,
        "unmatched": [],
    }

    if "restaurants" not in workbook.sheetnames:
        workbook.close()
        return stats

    sheet = workbook["restaurants"]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(cell) for cell in header_row]
    cols = {header: index for index, header in enumerate(headers) if header}

    required = {"id", "name", "city", "description_de", "description_en", "description_es"}
    missing = required - set(cols)
    if missing:
        workbook.close()
        raise ValueError(f"Spalten fehlen in {workbook_path.name}: {missing}")

    venue_col = cols.get("venue_type")

    for row_index in range(2, sheet.max_row + 1):
        restaurant_id = sheet.cell(row=row_index, column=cols["id"] + 1).value
        if not restaurant_id:
            continue
        restaurant_id = str(restaurant_id).strip()

        de_cell = sheet.cell(row=row_index, column=cols["description_de"] + 1)
        description_de = str(de_cell.value or "").strip()
        name = sheet.cell(row=row_index, column=cols["name"] + 1).value
        city = str(sheet.cell(row=row_index, column=cols["city"] + 1).value or "").strip()
        venue_type = (
            sheet.cell(row=row_index, column=venue_col + 1).value if venue_col is not None else None
        )

        if is_krumcoffee(name) and city:
            copy = krumcoffee_copy(city)
            de_cell.value = copy["de"]
            sheet.cell(row=row_index, column=cols["description_en"] + 1, value=copy["en"])
            sheet.cell(row=row_index, column=cols["description_es"] + 1, value=copy["es"])
            stats["special_copy"] += 1
            stats["en_set"] += 1
            stats["es_set"] += 1
            continue

        if restaurant_id in SEED_DE_FIXES:
            description_de = SEED_DE_FIXES[restaurant_id]
            de_cell.value = description_de
            stats["seed_de_fixed"] += 1

        needs_food_copy = not description_de or is_audit_style_description(description_de)
        if needs_food_copy and city:
            category = classify_venue_category(venue_type, name)
            copy = build_category_copy(category, city)
            de_cell.value = copy["de"]
            sheet.cell(row=row_index, column=cols["description_en"] + 1, value=copy["en"])
            sheet.cell(row=row_index, column=cols["description_es"] + 1, value=copy["es"])
            if description_de and is_audit_style_description(description_de):
                stats["audit_replaced"] += 1
            stats["en_set"] += 1
            stats["es_set"] += 1
            continue

        if not description_de:
            continue

        translated = translate_from_german(description_de)
        if not translated:
            stats["unmatched"].append((restaurant_id, description_de[:60]))
            continue

        en_text, es_text = translated
        sheet.cell(row=row_index, column=cols["description_en"] + 1, value=en_text)
        sheet.cell(row=row_index, column=cols["description_es"] + 1, value=es_text)
        stats["en_set"] += 1
        stats["es_set"] += 1

    workbook.save(workbook_path)
    workbook.close()
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Vereinheitlicht DE/EN/ES-Beschreibungen in den Excel-Workbooks."
    )
    parser.add_argument(
        "-t",
        "--target",
        type=Path,
        action="append",
        dest="targets",
        help="Ziel-Workbook (mehrfach moeglich)",
    )
    args = parser.parse_args()

    targets = args.targets or DEFAULT_TARGETS
    for target in targets:
        if not target.exists():
            print(f"Warnung: Ziel nicht gefunden, uebersprungen: {target}")
            continue

        stats = apply_to_workbook(target)
        print(f"\n{target.name}")
        print(f"  DE-Seed-Texte ersetzt: {stats['seed_de_fixed']}")
        print(f"  Audit-Texte ersetzt: {stats['audit_replaced']}")
        print(f"  Spezial-Texte (z. B. KrümCoffee): {stats['special_copy']}")
        print(f"  description_en gesetzt: {stats['en_set']}")
        print(f"  description_es gesetzt: {stats['es_set']}")
        if stats["unmatched"]:
            print(f"  Ohne Template-Treffer: {len(stats['unmatched'])}")
            for restaurant_id, snippet in stats["unmatched"]:
                print(f"    {restaurant_id}: {snippet}")


if __name__ == "__main__":
    main()
